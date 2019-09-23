from openpyxl import  load_workbook
from Common import read_path


class DoExcel():
    def __init__(self,filename,sheet):
        self.filename=filename
        self.sheet=sheet

    def get_lilte(self):
        wb=load_workbook(self.filename)
        sheet=wb[self.sheet]
        get_lilte=[]
        for i  in range(1,sheet.max_column+1):
            get_lilte.append(sheet.cell(1,i).value)
        return get_lilte

    def do_excel(self):
        wb = load_workbook(self.filename)
        sheet = wb[self.sheet]
        get_lilte = self.get_lilte()
        all_data=[]
        for j  in range(2,sheet.max_row+1):
            dict_data={}
            for i  in range(1,sheet.max_column+1):
                dict_data[get_lilte[i-1]]=sheet.cell(j,i).value
            all_data.append(dict_data)
        return all_data

    def write_back(self,row,col,new_value):
        wb = load_workbook(self.filename)
        sheet = wb[self.sheet]
        sheet.cell(row,col).value=new_value
        wb.save(self.filename)

